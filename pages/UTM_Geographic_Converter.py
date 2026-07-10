# Excel Coordinate Converter
# Input rows may contain either:
#   - WGS84 Geographic: lat / long
#   - WGS84 UTM Zone 17N: E / N
#
# Template displays "WGS84 UTM 17T" because Google Earth may show the latitude band as 17T.
# For pyproj, conversion is handled as WGS84 / UTM Zone 17N = EPSG:32617.
#
# This page fills the missing coordinate system:
#   - If lat/long is provided, it fills E/N.
#   - If E/N is provided, it fills lat/long.
#
# No NAD27, no NAD83, no NTv2 grids, no GTAA survey transformation.

from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from pyproj import Transformer

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# -------------------------------------------------------------------
# Streamlit page
# -------------------------------------------------------------------
st.set_page_config(page_title="Excel Coordinate Converter", page_icon="🔁")
st.title("Excel Coordinate Converter")

st.caption(
    """  
    Input coordinates to the Excel template provided.
    Rows may contain either **WGS84 Geographic** (`lat`, `long`) or **WGS84 UTM 17T** (`E`, `N`).  
    Output includes `feature_name`,`lat`, `long`,`E` and `N` (the missing pair is filled out).

    """
)


# -------------------------------------------------------------------
# Template download - same format as Excel_to_KMZ
# -------------------------------------------------------------------
def build_excel_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Group headers
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")
    ws["D1"] = "WGS84 Geographic"
    ws["F1"] = "WGS84 UTM 17T"

    # Column headers - E before N
    headers = [
        "folder",
        "subfolder",
        "feature_name",
        "lat",
        "long",
        "E",
        "N",
        "elevation (optional)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx).value = header

    # Formatting
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    geo_fill = PatternFill("solid", fgColor="EDEDED")
    utm_fill = PatternFill("solid", fgColor="EDEDED")
    header_fill = PatternFill("solid", fgColor="EDEDED")

    for cell_ref, fill in [("D1", geo_fill), ("F1", utm_fill)]:
        cell = ws[cell_ref]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Column widths
    widths = {
        "A": 18,  # folder
        "B": 18,  # subfolder
        "C": 26,  # feature_name
        "D": 16,  # lat
        "E": 16,  # long
        "F": 16,  # E
        "G": 16,  # N
        "H": 22,  # elevation optional
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # Number formats for user-entry area
    for row in range(3, 1003):
        ws[f"D{row}"].number_format = "0.000000000"  # lat
        ws[f"E{row}"].number_format = "0.000000000"  # long
        ws[f"F{row}"].number_format = "0.000"        # E
        ws[f"G{row}"].number_format = "0.000"        # N
        ws[f"H{row}"].number_format = "0.00"         # elevation

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


st.download_button(
    "Download Excel template",
    data=build_excel_template(),
    file_name="Excel_Coordinate_Converter_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


# -------------------------------------------------------------------
# File upload
# -------------------------------------------------------------------
up = st.file_uploader(
    "Upload Excel template",
    type=["xlsx", "xls"],
)


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [
        str(c).strip().lower().replace("\n", " ").replace(" ", "_")
        for c in df.columns
    ]
    return df


def pick(df: pd.DataFrame, options) -> str | None:
    for o in options:
        if o in df.columns:
            return o
    return None


def read_excel_template_aware(file_bytes: bytes) -> pd.DataFrame:
    """
    Current template uses row 2 as actual headers, so first try header=1.
    Fallback to header=0 for older/simple templates.
    """
    for header_row in [1, 0]:
        temp = pd.read_excel(BytesIO(file_bytes), header=header_row)
        temp = normalize_columns(temp)

        has_ll = (
            pick(temp, ["lat", "latitude"]) is not None
            and pick(temp, ["long", "lon", "longitude"]) is not None
        )
        has_utm = (
            pick(temp, ["e", "easting", "utm_e", "utm_easting", "x"]) is not None
            and pick(temp, ["n", "northing", "utm_n", "utm_northing", "y"]) is not None
        )

        if has_ll or has_utm:
            return temp

    return normalize_columns(pd.read_excel(BytesIO(file_bytes), header=1))


def invalid_ll(long_series: pd.Series, lat_series: pd.Series) -> pd.Series:
    long_num = pd.to_numeric(long_series, errors="coerce")
    lat_num = pd.to_numeric(lat_series, errors="coerce")

    bad = long_num.isna() | lat_num.isna()
    bad |= ~long_num.between(-180.0, 180.0)
    bad |= ~lat_num.between(-90.0, 90.0)

    return bad


def invalid_wgs84_utm17(e_series: pd.Series, n_series: pd.Series) -> pd.Series:
    e_num = pd.to_numeric(e_series, errors="coerce")
    n_num = pd.to_numeric(n_series, errors="coerce")

    bad = e_num.isna() | n_num.isna()
    bad |= ~e_num.between(100_000, 900_000)
    bad |= ~n_num.between(0, 10_000_000)

    return bad


def transformer_wgs84_geo_to_utm17() -> Transformer:
    """
    WGS84 Geographic degrees -> WGS84 / UTM Zone 17N meters.
    EPSG:4326 -> EPSG:32617.
    """
    return Transformer.from_crs("EPSG:4326", "EPSG:32617", always_xy=True)


def transformer_wgs84_utm17_to_geo() -> Transformer:
    """
    WGS84 / UTM Zone 17N meters -> WGS84 Geographic degrees.
    EPSG:32617 -> EPSG:4326.
    """
    return Transformer.from_crs("EPSG:32617", "EPSG:4326", always_xy=True)


def blank_if_nan(value):
    if pd.isna(value):
        return None
    return value


def build_output_excel(out: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Converted"

    # Group headers
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")
    ws["D1"] = "WGS84 Geographic"
    ws["F1"] = "WGS84 UTM 17T"

    headers = [
        "folder",
        "subfolder",
        "feature_name",
        "lat",
        "long",
        "E",
        "N",
        "elevation (optional)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx).value = header

    # Formatting
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    geo_fill = PatternFill("solid", fgColor="EDEDED")
    utm_fill = PatternFill("solid", fgColor="EDEDED")
    header_fill = PatternFill("solid", fgColor="EDEDED")

    for cell_ref, fill in [("D1", geo_fill), ("F1", utm_fill)]:
        cell = ws[cell_ref]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data
    for row_idx, record in enumerate(out.itertuples(index=False), start=3):
        ws.cell(row=row_idx, column=1).value = blank_if_nan(record.folder)
        ws.cell(row=row_idx, column=2).value = blank_if_nan(record.subfolder)
        ws.cell(row=row_idx, column=3).value = blank_if_nan(record.feature_name)
        ws.cell(row=row_idx, column=4).value = blank_if_nan(record.lat)
        ws.cell(row=row_idx, column=5).value = blank_if_nan(record.long)
        ws.cell(row=row_idx, column=6).value = blank_if_nan(record.E)
        ws.cell(row=row_idx, column=7).value = blank_if_nan(record.N)
        ws.cell(row=row_idx, column=8).value = blank_if_nan(record.elevation)

    # Column widths
    widths = {
        "A": 18,
        "B": 18,
        "C": 26,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # Number formats
    for row in range(3, ws.max_row + 1):
        ws[f"D{row}"].number_format = "0.000000000"
        ws[f"E{row}"].number_format = "0.000000000"
        ws[f"F{row}"].number_format = "0.000"
        ws[f"G{row}"].number_format = "0.000"
        ws[f"H{row}"].number_format = "0.00"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# -------------------------------------------------------------------
# Session persistence
# -------------------------------------------------------------------
for key in ("xlsx_bytes", "base_name", "summary"):
    if key not in st.session_state:
        st.session_state[key] = None

convert_clicked = st.button("Convert coordinates") if up else False


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------
if convert_clicked and up:
    try:
        df0 = read_excel_template_aware(up.getvalue())

        if df0.empty:
            st.error("No rows found.")
            st.stop()

        df = df0.copy()

        col_folder = pick(df, ["folder", "group", "layer"])
        col_subfolder = pick(df, ["subfolder", "sub_folder"])
        col_name = pick(df, ["feature_name", "name", "label", "id", "title"])
        col_lat = pick(df, ["lat", "latitude"])
        col_long = pick(df, ["long", "lon", "longitude"])
        col_e = pick(df, ["e", "easting", "utm_e", "utm_easting", "x"])
        col_n = pick(df, ["n", "northing", "utm_n", "utm_northing", "y"])
        col_z = pick(
            df,
            [
                "elevation",
                "elevation_(optional)",
                "elevation_optional",
                "elev",
                "z",
                "altitude",
                "height",
                "elevation_m",
            ],
        )

        if not ((col_lat and col_long) or (col_e and col_n)):
            st.error("Missing coordinate columns. Provide either lat/long or E/N.")
            st.stop()

        keep = [
            c for c in [
                col_folder,
                col_subfolder,
                col_name,
                col_lat,
                col_long,
                col_e,
                col_n,
                col_z,
            ]
            if c
        ]

        df = df[keep].copy()

        # Remove fully blank coordinate rows
        important_cols = [c for c in [col_lat, col_long, col_e, col_n] if c]
        df = df[~df[important_cols].isna().all(axis=1)].copy()

        if df.empty:
            st.error("No filled coordinate rows found.")
            st.stop()

        # Numeric coercion
        for c in [col_lat, col_long, col_e, col_n, col_z]:
            if c and c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        # Output frame, same columns as template
        out = pd.DataFrame(index=df.index)
        out["folder"] = df[col_folder] if col_folder and col_folder in df.columns else ""
        out["subfolder"] = df[col_subfolder] if col_subfolder and col_subfolder in df.columns else ""
        out["feature_name"] = df[col_name] if col_name and col_name in df.columns else ""

        out["lat"] = df[col_lat] if col_lat and col_lat in df.columns else np.nan
        out["long"] = df[col_long] if col_long and col_long in df.columns else np.nan
        out["E"] = df[col_e] if col_e and col_e in df.columns else np.nan
        out["N"] = df[col_n] if col_n and col_n in df.columns else np.nan
        out["elevation"] = df[col_z] if col_z and col_z in df.columns else np.nan

        ll_valid = ~invalid_ll(out["long"], out["lat"])
        utm_valid = ~invalid_wgs84_utm17(out["E"], out["N"])

        # Rows with valid lat/long and missing/invalid E/N: fill E/N
        m_geo_to_utm = ll_valid & ~utm_valid

        if m_geo_to_utm.any():
            tr_geo_to_utm = transformer_wgs84_geo_to_utm17()
            e_vals, n_vals = tr_geo_to_utm.transform(
                out.loc[m_geo_to_utm, "long"].to_numpy(),
                out.loc[m_geo_to_utm, "lat"].to_numpy(),
            )

            out.loc[m_geo_to_utm, "E"] = e_vals
            out.loc[m_geo_to_utm, "N"] = n_vals

        # Recalculate UTM validity after filling
        utm_valid_after = ~invalid_wgs84_utm17(out["E"], out["N"])

        # Rows with valid E/N and missing/invalid lat/long: fill lat/long
        m_utm_to_geo = utm_valid_after & ~ll_valid

        if m_utm_to_geo.any():
            tr_utm_to_geo = transformer_wgs84_utm17_to_geo()
            long_vals, lat_vals = tr_utm_to_geo.transform(
                out.loc[m_utm_to_geo, "E"].to_numpy(),
                out.loc[m_utm_to_geo, "N"].to_numpy(),
            )

            out.loc[m_utm_to_geo, "long"] = long_vals
            out.loc[m_utm_to_geo, "lat"] = lat_vals

        # Final validation
        final_ll_valid = ~invalid_ll(out["long"], out["lat"])
        final_utm_valid = ~invalid_wgs84_utm17(out["E"], out["N"])
        bad_final = ~(final_ll_valid & final_utm_valid)

        if bad_final.any():
            st.warning(
                f"{bad_final.sum()} row(s) could not be completed because neither valid lat/long nor valid E/N was available."
            )

        # Keep valid completed rows only
        out = out[~bad_final].copy()

        if out.empty:
            st.error("No valid rows were available after conversion.")
            st.stop()

        # Round output
        out["lat"] = pd.to_numeric(out["lat"], errors="coerce").round(9)
        out["long"] = pd.to_numeric(out["long"], errors="coerce").round(9)
        out["E"] = pd.to_numeric(out["E"], errors="coerce").round(3)
        out["N"] = pd.to_numeric(out["N"], errors="coerce").round(3)
        out["elevation"] = pd.to_numeric(out["elevation"], errors="coerce").round(2)

        base = Path(up.name).stem
        xlsx_bytes = build_output_excel(out)

        st.session_state.xlsx_bytes = xlsx_bytes
        st.session_state.base_name = base
        st.session_state.summary = (
            f"Completed {len(out)} row(s). "
            f"Filled E/N from lat/long on {int(m_geo_to_utm.sum())} row(s). "
            f"Filled lat/long from E/N on {int(m_utm_to_geo.sum())} row(s)."
        )

    except Exception as e:
        st.error("Coordinate conversion failed.")
        st.exception(e)


# -------------------------------------------------------------------
# Downloads
# -------------------------------------------------------------------
if st.session_state.xlsx_bytes:
    st.success(st.session_state.summary or "Coordinate conversion complete.")

    st.download_button(
        "Download completed Excel",
        data=st.session_state.xlsx_bytes,
        file_name=f"{st.session_state.base_name}_completed.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx",
    )

    def _queue_refresh():
        for k in ("xlsx_bytes", "base_name", "summary"):
            st.session_state.pop(k, None)
        st.session_state["_do_rerun"] = True

    st.button("Refresh / New conversion", on_click=_queue_refresh, type="secondary")


if st.session_state.get("_do_rerun"):
    st.session_state.pop("_do_rerun", None)
    st.rerun()
