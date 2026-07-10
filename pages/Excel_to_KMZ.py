# Excel → KMZ
# Input rows may contain either:
#   - WGS84 Geographic: lat / long
#   - WGS84 UTM Zone 17N: E / N
#
# Template displays "WGS84 UTM 17T" because Google Earth may show the latitude band as 17T.
# For pyproj, conversion is handled as WGS84 / UTM Zone 17N = EPSG:32617.
#
# No NAD27, no NAD83, no NTv2 grids, no GTAA survey transformation.

import zipfile
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

import numpy as np
import pandas as pd
import streamlit as st
from pyproj import Transformer

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# -------------------------------------------------------------------
# Streamlit page
# -------------------------------------------------------------------
st.set_page_config(page_title="Excel to KMZ", page_icon="🌎")
st.title("Excel to KMZ")

st.caption(
    """
    Input coordinates to the Excel template provided.  
    Rows may contain either **WGS84 Geographic** (`lat`, `long`) or **WGS84 UTM 17T** (`E`, `N`).  
    Folder and subfolder information are optional and can be used to nest features in Google Earth. Elevation is optional.  

    **Note:** Visually confirm the geographic placement of points in the new file.
    """
)


# -------------------------------------------------------------------
# Template download - generated in code
# -------------------------------------------------------------------
def build_excel_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")
    ws["D1"] = "WGS84 Geographic"
    ws["F1"] = "WGS84 UTM 17T"

    headers = [
        "folder",
        "subfolder",
        "feature_name",
        "lat",
        "long",
        "E",
        "N",
        "elevation (optional)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx).value = header

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    geo_fill = PatternFill("solid", fgColor="EDEDED")
    utm_fill = PatternFill("solid", fgColor="EDEDED")
    header_fill = PatternFill("solid", fgColor="EDEDED")

    for cell_ref, fill in [("D1", geo_fill), ("F1", utm_fill)]:
        cell = ws[cell_ref]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = {
        "A": 18,
        "B": 18,
        "C": 26,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for row in range(3, 1003):
        ws[f"D{row}"].number_format = "0.000000000"
        ws[f"E{row}"].number_format = "0.000000000"
        ws[f"F{row}"].number_format = "0.000"
        ws[f"G{row}"].number_format = "0.000"
        ws[f"H{row}"].number_format = "0.00"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


st.download_button(
    "Download Excel template",
    data=build_excel_template(),
    file_name="Excel_to_KMZ_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


# -------------------------------------------------------------------
# File upload
# -------------------------------------------------------------------
up = st.file_uploader(
    "Upload Excel template",
    type=["xlsx", "xls"],
)


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [
        str(c).strip().lower().replace("\n", " ").replace(" ", "_")
        for c in df.columns
    ]
    return df


def pick(df: pd.DataFrame, options) -> str | None:
    for o in options:
        if o in df.columns:
            return o
    return None


def read_excel_template_aware(file_bytes: bytes) -> pd.DataFrame:
    for header_row in [1, 0]:
        temp = pd.read_excel(BytesIO(file_bytes), header=header_row)
        temp = normalize_columns(temp)

        has_name = pick(temp, ["feature_name", "name", "label", "id", "title"]) is not None
        has_ll = (
            pick(temp, ["lat", "latitude"]) is not None
            and pick(temp, ["long", "lon", "longitude"]) is not None
        )
        has_utm = (
            pick(temp, ["e", "easting", "utm_e", "utm_easting", "x"]) is not None
            and pick(temp, ["n", "northing", "utm_n", "utm_northing", "y"]) is not None
        )

        if has_name and (has_ll or has_utm):
            return temp

    return normalize_columns(pd.read_excel(BytesIO(file_bytes), header=1))


def invalid_ll(long_series: pd.Series, lat_series: pd.Series) -> pd.Series:
    long_num = pd.to_numeric(long_series, errors="coerce")
    lat_num = pd.to_numeric(lat_series, errors="coerce")

    bad = long_num.isna() | lat_num.isna()
    bad |= ~long_num.between(-180.0, 180.0)
    bad |= ~lat_num.between(-90.0, 90.0)
    return bad


def invalid_wgs84_utm17(e_series: pd.Series, n_series: pd.Series) -> pd.Series:
    e_num = pd.to_numeric(e_series, errors="coerce")
    n_num = pd.to_numeric(n_series, errors="coerce")

    bad = e_num.isna() | n_num.isna()
    bad |= ~e_num.between(100_000, 900_000)
    bad |= ~n_num.between(0, 10_000_000)
    return bad


def transformer_wgs84_utm17_to_wgs84_geo() -> Transformer:
    return Transformer.from_crs("EPSG:32617", "EPSG:4326", always_xy=True)


def clean_folder_value(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def add_note(df: pd.DataFrame, mask: pd.Series, note: str) -> None:
    if not mask.any():
        return

    existing = df.loc[mask, "conversion_note"].fillna("").astype(str).str.strip()
    df.loc[mask, "conversion_note"] = np.where(
        existing == "",
        note,
        existing + "; " + note,
    )


def build_kmz_points(doc_name: str, rows: pd.DataFrame) -> bytes:
    def kml_header(name):
        return (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<kml xmlns="http://www.opengis.net/kml/2.2">\n'
            f'  <Document><name>{escape(name)}</name>\n'
        )

    def kml_footer():
        return "  </Document>\n</kml>\n"

    def kml_folder(name, indent="    "):
        return f'{indent}<Folder><name>{escape(name)}</name>\n'

    def kml_folder_end(indent="    "):
        return f"{indent}</Folder>\n"

    def pm_point(name, long_val, lat_val, elev=None, indent="      "):
        coord = f"{long_val:.9f},{lat_val:.9f}"
        if elev is not None and np.isfinite(elev):
            coord += f",{float(elev):.2f}"

        return (
            f"{indent}<Placemark>\n"
            f"{indent}  <name>{escape(str(name))}</name>\n"
            f"{indent}  <Point><coordinates>{coord}</coordinates></Point>\n"
            f"{indent}</Placemark>\n"
        )

    def write_points(kml_buf, point_rows, indent="      "):
        for _, r in point_rows.iterrows():
            kml_buf.write(
                pm_point(
                    r["feature_name"],
                    float(r["long"]),
                    float(r["lat"]),
                    r.get("elevation", np.nan),
                    indent=indent,
                ).encode("utf-8")
            )

    work = rows.copy()
    work["_folder"] = work["folder"].map(clean_folder_value) if "folder" in work.columns else ""
    work["_subfolder"] = work["subfolder"].map(clean_folder_value) if "subfolder" in work.columns else ""

    kml = BytesIO()
    kml.write(kml_header(doc_name).encode("utf-8"))

    doc_level = work[(work["_folder"] == "") & (work["_subfolder"] == "")]
    write_points(kml, doc_level, indent="    ")

    subfolder_only = work[(work["_folder"] == "") & (work["_subfolder"] != "")]
    for subfolder_name, sg in subfolder_only.groupby("_subfolder", sort=False):
        kml.write(kml_folder(subfolder_name, indent="    ").encode("utf-8"))
        write_points(kml, sg, indent="      ")
        kml.write(kml_folder_end(indent="    ").encode("utf-8"))

    with_folder = work[work["_folder"] != ""]
    for folder_name, fg in with_folder.groupby("_folder", sort=False):
        kml.write(kml_folder(folder_name, indent="    ").encode("utf-8"))

        direct = fg[fg["_subfolder"] == ""]
        write_points(kml, direct, indent="      ")

        nested = fg[fg["_subfolder"] != ""]
        for subfolder_name, sg in nested.groupby("_subfolder", sort=False):
            kml.write(kml_folder(subfolder_name, indent="      ").encode("utf-8"))
            write_points(kml, sg, indent="        ")
            kml.write(kml_folder_end(indent="      ").encode("utf-8"))

        kml.write(kml_folder_end(indent="    ").encode("utf-8"))

    kml.write(kml_footer().encode("utf-8"))
    kml.seek(0)

    kmz = BytesIO()
    with zipfile.ZipFile(kmz, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("doc.kml", kml.getvalue())

    kmz.seek(0)
    return kmz.getvalue()


# -------------------------------------------------------------------
# Session persistence
# -------------------------------------------------------------------
for key in ("kmz_bytes", "validation_xlsx", "base_name"):
    if key not in st.session_state:
        st.session_state[key] = None

convert_clicked = st.button("Convert") if up else False


# -------------------------------------------------------------------
# Main conversion
# -------------------------------------------------------------------
if convert_clicked and up:
    try:
        df0 = read_excel_template_aware(up.getvalue())

        if df0.empty:
            st.error("No rows found.")
        else:
            df = df0.copy()

            col_folder = pick(df, ["folder", "group", "layer"])
            col_subfolder = pick(df, ["subfolder", "sub_folder"])
            col_name = pick(df, ["feature_name", "name", "label", "id", "title"])
            col_lat = pick(df, ["lat", "latitude"])
            col_long = pick(df, ["long", "lon", "longitude"])
            col_e = pick(df, ["e", "easting", "utm_e", "utm_easting", "x"])
            col_n = pick(df, ["n", "northing", "utm_n", "utm_northing", "y"])
            col_z = pick(
                df,
                [
                    "elevation",
                    "elevation_(optional)",
                    "elevation_optional",
                    "elev",
                    "z",
                    "altitude",
                    "height",
                    "elevation_m",
                ],
            )

            if col_name is None:
                st.error("Missing required column: feature_name")
            elif not ((col_lat and col_long) or (col_e and col_n)):
                st.error("Missing coordinate columns. Provide either lat/long or E/N.")
            else:
                keep = [
                    c for c in [
                        col_folder,
                        col_subfolder,
                        col_name,
                        col_lat,
                        col_long,
                        col_e,
                        col_n,
                        col_z,
                    ]
                    if c
                ]

                df = df[keep].copy()

                important_cols = [c for c in [col_name, col_lat, col_long, col_e, col_n] if c]
                df = df[~df[important_cols].isna().all(axis=1)].copy()

                if df.empty:
                    st.error("No filled rows found.")
                    st.stop()

                for c in [col_lat, col_long, col_e, col_n, col_z]:
                    if c and c in df.columns:
                        df[c] = pd.to_numeric(df[c], errors="coerce")

                out = pd.DataFrame(index=df.index)
                out["folder"] = df[col_folder] if col_folder and col_folder in df.columns else ""
                out["subfolder"] = df[col_subfolder] if col_subfolder and col_subfolder in df.columns else ""

                out["feature_name"] = df[col_name].fillna("").astype(str).str.strip()
                blank_names = out["feature_name"] == ""
                for seq, idx in enumerate(out.index[blank_names], start=1):
                    out.loc[idx, "feature_name"] = f"Point {seq}"

                out["elevation"] = df[col_z] if col_z and col_z in df.columns else np.nan

                out["lat"] = df[col_lat] if col_lat and col_lat in df.columns else np.nan
                out["long"] = df[col_long] if col_long and col_long in df.columns else np.nan
                out["E"] = df[col_e] if col_e and col_e in df.columns else np.nan
                out["N"] = df[col_n] if col_n and col_n in df.columns else np.nan

                out["validation_status"] = ""
                out["input_type"] = ""
                out["conversion_note"] = ""

                # Validate provided inputs
                if col_lat and col_long:
                    m_latlong_provided = df[col_lat].notna() & df[col_long].notna()
                    m_latlong_valid = m_latlong_provided & ~invalid_ll(df[col_long], df[col_lat])
                    m_latlong_invalid = m_latlong_provided & invalid_ll(df[col_long], df[col_lat])
                else:
                    m_latlong_provided = pd.Series(False, index=df.index)
                    m_latlong_valid = pd.Series(False, index=df.index)
                    m_latlong_invalid = pd.Series(False, index=df.index)

                if col_e and col_n:
                    m_utm_provided = df[col_e].notna() & df[col_n].notna()
                    m_utm_valid = m_utm_provided & ~invalid_wgs84_utm17(df[col_e], df[col_n])
                    m_utm_invalid = m_utm_provided & invalid_wgs84_utm17(df[col_e], df[col_n])
                else:
                    m_utm_provided = pd.Series(False, index=df.index)
                    m_utm_valid = pd.Series(False, index=df.index)
                    m_utm_invalid = pd.Series(False, index=df.index)

                # A) Valid WGS84 geographic input: pass through directly
                out.loc[m_latlong_valid, "input_type"] = "wgs84_geographic"
                add_note(out, m_latlong_valid, "plotted_from_lat_long")

                # B) Rows with both valid lat/long and valid E/N: lat/long is used for KMZ
                m_both_valid = m_latlong_valid & m_utm_valid
                out.loc[m_both_valid, "input_type"] = "wgs84_geographic_preferred"
                add_note(out, m_both_valid, "both_lat_long_and_utm_present_lat_long_used")

                # C) Valid UTM, but no valid lat/long: convert E/N to WGS84 geographic
                m_utm_to_geo = m_utm_valid & ~m_latlong_valid

                if m_utm_to_geo.any():
                    tr = transformer_wgs84_utm17_to_wgs84_geo()
                    long_vals, lat_vals = tr.transform(
                        df.loc[m_utm_to_geo, col_e].to_numpy(),
                        df.loc[m_utm_to_geo, col_n].to_numpy(),
                    )

                    out.loc[m_utm_to_geo, "long"] = long_vals
                    out.loc[m_utm_to_geo, "lat"] = lat_vals
                    out.loc[m_utm_to_geo, "input_type"] = "wgs84_utm_17t"
                    add_note(out, m_utm_to_geo, "converted_from_wgs84_utm_zone_17n")

                # Input problems that should remain visible in Validation Excel
                add_note(out, m_latlong_invalid, "invalid_lat_long_failed_range_check")
                add_note(out, m_utm_invalid, "invalid_utm_e_n_failed_range_check")

                # Final coordinate validation for KMZ export
                bad_final = invalid_ll(out["long"], out["lat"])
                good_final = ~bad_final

                out.loc[good_final, "validation_status"] = "OK - included in KMZ"
                out.loc[bad_final, "validation_status"] = "ERROR - skipped from KMZ"
                add_note(out, bad_final, "missing_or_invalid_final_wgs84_coordinates")

                failed_utm_no_latlong = m_utm_invalid & ~m_latlong_valid
                if failed_utm_no_latlong.any():
                    st.warning(
                        f"{failed_utm_no_latlong.sum()} UTM row(s) failed WGS84 UTM 17T range checks and are shown in the Validation Excel."
                    )

                if bad_final.any():
                    st.warning(
                        f"{bad_final.sum()} row(s) had invalid or missing final WGS84 coordinates and are shown in the Validation Excel."
                    )

                # Round validation output, but keep invalid/error rows
                out["lat"] = pd.to_numeric(out["lat"], errors="coerce").round(9)
                out["long"] = pd.to_numeric(out["long"], errors="coerce").round(9)
                out["E"] = pd.to_numeric(out["E"], errors="coerce").round(3)
                out["N"] = pd.to_numeric(out["N"], errors="coerce").round(3)
                out["elevation"] = pd.to_numeric(out["elevation"], errors="coerce")

                base = Path(up.name).stem

                # KMZ only uses valid rows
                kmz_rows = out[out["validation_status"] == "OK - included in KMZ"].copy()

                kmz_bytes = None
                if not kmz_rows.empty:
                    kmz_bytes = build_kmz_points(
                        f"{base}",
                        kmz_rows[[
                            "folder",
                            "subfolder",
                            "feature_name",
                            "long",
                            "lat",
                            "elevation",
                        ]],
                    )
                else:
                    st.warning("No valid rows were available for KMZ export. Validation Excel only.")

                # Validation Excel includes valid + error rows
                valid_cols = [
                    "folder",
                    "subfolder",
                    "feature_name",
                    "lat",
                    "long",
                    "E",
                    "N",
                    "elevation",
                    "validation_status",
                    "input_type",
                    "conversion_note",
                ]

                out_valid = out[valid_cols].copy()

                xbuf = BytesIO()
                with pd.ExcelWriter(xbuf, engine="openpyxl") as xw:
                    readme = pd.DataFrame(
                        {
                            "Validation": [
                                "All valid KMZ export coordinates are WGS84 Geographic longitude/latitude.",
                                "Rows with valid lat/long are plotted directly.",
                                "Rows with valid E/N and missing/invalid lat/long are interpreted as WGS84 UTM 17T / UTM Zone 17N and converted to WGS84 Geographic.",
                                "Rows with invalid or missing final WGS84 coordinates are skipped from the KMZ but kept in this Validation Excel.",
                                "No NAD27, NAD83, NTv2 grid, or GTAA survey transformation is applied in this tool.",
                            ]
                        }
                    )
                    readme.to_excel(xw, index=False, sheet_name="README")
                    out_valid.to_excel(xw, index=False, sheet_name="Validation")

                xbuf.seek(0)

                st.session_state.kmz_bytes = kmz_bytes
                st.session_state.validation_xlsx = xbuf.getvalue()
                st.session_state.base_name = base

    except Exception as e:
        st.error("Conversion failed.")
        st.exception(e)


# -------------------------------------------------------------------
# Downloads
# -------------------------------------------------------------------
if st.session_state.validation_xlsx:
    if st.session_state.kmz_bytes:
        st.success("KMZ and Validation Excel are ready.")
    else:
        st.success("Validation Excel is ready. No valid KMZ rows were available.")

    if st.session_state.kmz_bytes:
        st.download_button(
            "Download KMZ",
            data=st.session_state.kmz_bytes,
            file_name=f"{st.session_state.base_name}.kmz",
            mime="application/vnd.google-earth.kmz",
            key="dl_kmz",
        )

    st.download_button(
        "Download Validation Excel",
        data=st.session_state.validation_xlsx,
        file_name=f"{st.session_state.base_name}_Validation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx",
    )

    def _queue_refresh():
        for k in ("kmz_bytes", "validation_xlsx", "base_name"):
            st.session_state.pop(k, None)
        st.session_state["_do_rerun"] = True

    st.button("Refresh / New conversion", on_click=_queue_refresh, type="secondary")


if st.session_state.get("_do_rerun"):
    st.session_state.pop("_do_rerun", None)
    st.rerun()
