# KMZ/KML → Excel
# Extracts native Google Earth WGS84 Geographic coordinates and adds WGS84 UTM 17T E/N.
#
# Output:
#   feature_name, vertex_index, lat, long, E, N, elevation (m)
#
# No NAD27, no NAD83, no NTv2 grids, no GTAA survey transformation.

import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from pyproj import Transformer

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# -------------------------------------------------------------------
# Streamlit page
# -------------------------------------------------------------------
st.set_page_config(page_title="KMZ Coordinates Extraction", page_icon="🧭")
st.title("KMZ Coordinates to Excel")

st.caption(
    """
    Upload a KMZ or KML file to extract the native **WGS84 Geographic** coordinates used by Google Earth.  
    Output includes `feature_name`, `vertex_index`,**WGS84 Geographic** `lat`, `long`, **WGS84 UTM 17T** `E`, `N`, and elevation when available.  
    """
)

up = st.file_uploader("Upload KMZ or KML", type=["kmz", "kml"])


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def transformer_wgs84_geo_to_utm17() -> Transformer:
    """
    WGS84 Geographic degrees -> WGS84 / UTM Zone 17N meters.
    Google Earth may display this as UTM 17T in the Toronto/Pearson area.
    """
    return Transformer.from_crs("EPSG:4326", "EPSG:32617", always_xy=True)


def parse_kml_bytes(kml_bytes: bytes) -> pd.DataFrame:
    """
    Parse KML bytes and return:
      feature_name, vertex_index, lat, long, elevation

    For single-point placemarks, vertex_index is left blank.
    For lines/polygons/multiple-coordinate placemarks, vertex_index is 1, 2, 3...
    """
    ns = {"kml": "http://www.opengis.net/kml/2.2"}
    root = ET.fromstring(kml_bytes)

    rows = []

    for pm in root.findall(".//kml:Placemark", ns):
        name_el = pm.find("kml:name", ns)
        name = name_el.text.strip() if name_el is not None and name_el.text else "Unnamed"

        coords_for_pm = []

        for ct in pm.findall(".//kml:coordinates", ns):
            text = (ct.text or "").strip()
            if not text:
                continue

            for c in text.split():
                parts = c.split(",")
                if len(parts) < 2:
                    continue

                try:
                    long_val = float(parts[0])
                    lat_val = float(parts[1])
                except ValueError:
                    continue

                elev = None
                if len(parts) > 2 and parts[2] != "":
                    try:
                        elev = float(parts[2])
                    except ValueError:
                        elev = None

                coords_for_pm.append(
                    {
                        "feature_name": name,
                        "lat": lat_val,
                        "long": long_val,
                        "elevation": elev,
                    }
                )

        multiple_vertices = len(coords_for_pm) > 1

        for idx, row in enumerate(coords_for_pm, start=1):
            row["vertex_index"] = idx if multiple_vertices else ""
            rows.append(row)

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df = df[[
        "feature_name",
        "vertex_index",
        "lat",
        "long",
        "elevation",
    ]].copy()

    df["lat"] = pd.to_numeric(df["lat"], errors="coerce")
    df["long"] = pd.to_numeric(df["long"], errors="coerce")
    df["elevation"] = pd.to_numeric(df["elevation"], errors="coerce")

    # Remove invalid geographic rows
    valid = (
        df["lat"].between(-90, 90)
        & df["long"].between(-180, 180)
        & df["lat"].notna()
        & df["long"].notna()
    )
    df = df[valid].copy()

    if df.empty:
        return df

    # Add WGS84 UTM 17T / Zone 17N E/N
    tr = transformer_wgs84_geo_to_utm17()
    e_vals, n_vals = tr.transform(
        df["long"].to_numpy(),
        df["lat"].to_numpy(),
    )

    df["E"] = pd.Series(e_vals, index=df.index, dtype="float64")
    df["N"] = pd.Series(n_vals, index=df.index, dtype="float64")

    # Round output values
    df["lat"] = df["lat"].round(9)
    df["long"] = df["long"].round(9)
    df["E"] = df["E"].round(3)
    df["N"] = df["N"].round(3)
    df["elevation"] = df["elevation"].round(2)

    # Final output order
    df = df[[
        "feature_name",
        "vertex_index",
        "lat",
        "long",
        "E",
        "N",
        "elevation",
    ]].copy()

    return df


def blank_if_nan(value):
    if pd.isna(value):
        return None
    return value


def build_output_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"

    # Group headers
    ws.merge_cells("C1:D1")
    ws.merge_cells("E1:F1")
    ws["C1"] = "WGS84 Geographic"
    ws["E1"] = "WGS84 UTM 17T"

    # Column headers
    headers = [
        "feature_name",
        "vertex_index",
        "lat",
        "long",
        "E",
        "N",
        "elevation (m)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx).value = header

    # Formatting similar to Excel_to_KMZ
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    geo_fill = PatternFill("solid", fgColor="EDEDED")
    utm_fill = PatternFill("solid", fgColor="EDEDED")
    header_fill = PatternFill("solid", fgColor="EDEDED")

    for cell_ref, fill in [("C1", geo_fill), ("E1", utm_fill)]:
        cell = ws[cell_ref]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows
    for row_idx, record in enumerate(df.itertuples(index=False), start=3):
        ws.cell(row=row_idx, column=1).value = record.feature_name
        ws.cell(row=row_idx, column=2).value = record.vertex_index
        ws.cell(row=row_idx, column=3).value = blank_if_nan(record.lat)
        ws.cell(row=row_idx, column=4).value = blank_if_nan(record.long)
        ws.cell(row=row_idx, column=5).value = blank_if_nan(record.E)
        ws.cell(row=row_idx, column=6).value = blank_if_nan(record.N)
        ws.cell(row=row_idx, column=7).value = blank_if_nan(record.elevation)

    # Column widths
    widths = {
        "A": 28,  # feature_name
        "B": 14,  # vertex_index
        "C": 16,  # lat
        "D": 16,  # long
        "E": 16,  # E
        "F": 16,  # N
        "G": 16,  # elevation
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # Number formats
    for row in range(3, ws.max_row + 1):
        ws[f"C{row}"].number_format = "0.000000000"
        ws[f"D{row}"].number_format = "0.000000000"
        ws[f"E{row}"].number_format = "0.000"
        ws[f"F{row}"].number_format = "0.000"
        ws[f"G{row}"].number_format = "0.00"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# -------------------------------------------------------------------
# Session persistence
# -------------------------------------------------------------------
for key in ("xlsx_bytes", "base_name"):
    if key not in st.session_state:
        st.session_state[key] = None

extract_clicked = st.button("Extract coordinates") if up else False


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------
if extract_clicked and up:
    try:
        # Read KML bytes from KMZ or KML
        if up.name.lower().endswith(".kmz"):
            with zipfile.ZipFile(up) as z:
                kml_files = [n for n in z.namelist() if n.lower().endswith(".kml")]

                if not kml_files:
                    st.error("No KML file found inside the KMZ.")
                    st.stop()

                kml_name = "doc.kml" if "doc.kml" in kml_files else kml_files[0]
                kml_bytes = z.read(kml_name)
        else:
            kml_bytes = up.read()

        df = parse_kml_bytes(kml_bytes)

        if df.empty:
            st.error("No valid coordinates found in the KMZ/KML.")
            st.stop()

        base = Path(up.name).stem
        xlsx_bytes = build_output_excel(df)

        st.session_state.xlsx_bytes = xlsx_bytes
        st.session_state.base_name = base

    except Exception as e:
        st.error("KMZ/KML coordinate extraction failed.")
        st.exception(e)


# -------------------------------------------------------------------
# Downloads
# -------------------------------------------------------------------
if st.session_state.xlsx_bytes:
    st.success("Coordinates Excel is ready.")

    st.download_button(
        "Download Excel",
        data=st.session_state.xlsx_bytes,
        file_name=f"{st.session_state.base_name}_coordinates.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_xlsx",
    )

    def _queue_refresh():
        for k in ("xlsx_bytes", "base_name"):
            st.session_state.pop(k, None)
        st.session_state["_do_rerun"] = True

    st.button("Refresh / New extraction", on_click=_queue_refresh, type="secondary")


if st.session_state.get("_do_rerun"):
    st.session_state.pop("_do_rerun", None)
    st.rerun()
